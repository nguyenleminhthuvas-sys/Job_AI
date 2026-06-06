#!/usr/bin/env python3
import fitz  # PyMuPDF
import openpyxl
import sys

BASE = '/Users/mac/Documents/AI Antigravity/SKILL/phan-tich-bctc/'

# ── Extract PDF ──
print("=== EXTRACTING PDF ===", flush=True)
doc = fitz.open(BASE + '20260414_NVLG_AR2025_VN.pdf')
print(f"PDF pages: {len(doc)}", flush=True)

pdf_text = []
for i, page in enumerate(doc):
    text = page.get_text()
    if text.strip():
        pdf_text.append(f"\n--- PAGE {i+1} ---\n{text}")

full_pdf = ''.join(pdf_text)
print(f"PDF extracted: {len(full_pdf)} chars, {full_pdf.count(chr(10))} lines", flush=True)

with open(BASE + 'NVL_AR2025_extracted.txt', 'w', encoding='utf-8') as f:
    f.write(full_pdf)
print("PDF saved to NVL_AR2025_extracted.txt", flush=True)

# ── Extract XLSX ──
print("\n=== EXTRACTING XLSX ===", flush=True)
wb = openpyxl.load_workbook(BASE + 'SSI_NVL_Financial_statement_Income_Statement_02062026 (1).xlsx')
print(f"Sheets: {wb.sheetnames}", flush=True)

xlsx_text = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    xlsx_text.append(f"\n=== SHEET: {sheet_name} ===\n")
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            row_str = '\t'.join([str(c) if c is not None else '' for c in row])
            xlsx_text.append(row_str + '\n')

full_xlsx = ''.join(xlsx_text)
print(f"XLSX extracted: {len(full_xlsx)} chars", flush=True)

with open(BASE + 'NVL_financials_extracted.txt', 'w', encoding='utf-8') as f:
    f.write(full_xlsx)
print("XLSX saved to NVL_financials_extracted.txt", flush=True)
print("DONE", flush=True)
