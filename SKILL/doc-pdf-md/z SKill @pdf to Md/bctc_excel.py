"""
SKILL: BCTC to Excel
Chuyển đổi PDF Báo cáo tài chính → Markdown + Excel (.xlsx)
Engine: docling (AI TableFormer) + openpyxl
Author: Nguyen Tung Chi — nguyentungchi.com
"""

import argparse
import os
import sys
import glob


# ──────────────────────────────────────────────────────────
# HELPER: Định dạng Excel đẹp
# ──────────────────────────────────────────────────────────
def apply_excel_style(ws, header_row_count=1):
    """Áp dụng style cho worksheet: header màu xanh đậm, border, freeze panes."""
    try:
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )

        # Màu header: xanh navy đậm
        HEADER_BG  = "1F3864"
        HEADER_FG  = "FFFFFF"
        ROW_ALT_BG = "EEF2F8"  # xanh nhạt xen kẽ

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

                if row_idx <= header_row_count:
                    # Header rows
                    cell.font = Font(name="Calibri", bold=True,
                                     color=HEADER_FG, size=11)
                    cell.fill = PatternFill(
                        fill_type="solid", fgColor=HEADER_BG
                    )
                else:
                    cell.font = Font(name="Calibri", size=10)
                    # Xen kẽ màu hàng
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(
                            fill_type="solid", fgColor=ROW_ALT_BG
                        )
                    # Tự động format số kế toán
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

        # Freeze header
        ws.freeze_panes = f"A{header_row_count + 1}"

        # Auto-fit column width (giới hạn max 50)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    except ImportError:
        pass  # Nếu không có openpyxl styles thì bỏ qua style


# ──────────────────────────────────────────────────────────
# CORE: PDF BCTC → Markdown + Excel
# ──────────────────────────────────────────────────────────
def convert_bctc(pdf_path: str, out_md: str, out_xlsx: str):
    """
    Bóc tách BCTC PDF → cả Markdown lẫn Excel.

    Args:
        pdf_path : đường dẫn file PDF
        out_md   : đường dẫn file .md xuất ra
        out_xlsx : đường dẫn file .xlsx xuất ra
    """
    # ── Bước 1: Dùng docling phân tích PDF ──
    print("🔄 Đang phân tích PDF bằng AI TableFormer (docling)...")
    try:
        from docling.document_converter import DocumentConverter
    except ImportError:
        print("❌ Lỗi: Cần cài docling  →  pip3 install docling")
        sys.exit(1)

    converter = DocumentConverter()
    try:
        result = converter.convert(pdf_path)
    except Exception as e:
        print(f"❌ Lỗi khi đọc PDF: {e}")
        sys.exit(1)

    doc = result.document

    # ── Bước 2: Xuất Markdown ──
    os.makedirs(os.path.dirname(out_md) or ".", exist_ok=True)
    md_text = doc.export_to_markdown()
    with open(out_md, "w", encoding="utf-8") as f:
        f.write(md_text)
    md_lines = md_text.count("\n")
    print(f"✅ Markdown: {out_md}  ({md_lines:,} dòng)")

    # ── Bước 3: Xuất Excel ──
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        print("❌ Lỗi: Cần cài openpyxl  →  pip3 install openpyxl")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định

    # Lấy tất cả bảng biểu từ docling
    tables = [item for item in doc.tables] if hasattr(doc, "tables") else []

    # Fallback: nếu docling không expose tables trực tiếp,
    # parse bảng từ Markdown
    if not tables:
        print("  ℹ️  Bóc bảng từ Markdown (fallback)...")
        _write_excel_from_markdown(wb, md_text)
    else:
        # Ghi từng bảng vào sheet riêng
        for i, table in enumerate(tables, start=1):
            sheet_name = f"Bảng {i}"
            ws = wb.create_sheet(title=sheet_name)

            try:
                # Lấy dữ liệu từ TableItem của docling
                df = table.export_to_dataframe()
                # Ghi header
                ws.append(list(df.columns))
                # Ghi dữ liệu
                for _, row in df.iterrows():
                    ws.append(
                        [_parse_number(v) for v in row.values]
                    )
            except Exception:
                # Fallback: export dạng list of lists
                try:
                    rows = table.export_to_list()
                    for row in rows:
                        ws.append([_parse_number(c) for c in row])
                except Exception as e2:
                    ws.append([f"[Lỗi đọc bảng: {e2}]"])

            apply_excel_style(ws, header_row_count=1)

    # Sheet tổng hợp: toàn bộ nội dung MD (text only)
    ws_raw = wb.create_sheet(title="📄 Full Text")
    ws_raw.append(["Nội dung Markdown đầy đủ"])
    ws_raw.append([""])
    for line in md_text.split("\n")[:2000]:  # Giới hạn 2000 dòng
        ws_raw.append([line])
    ws_raw.column_dimensions["A"].width = 120

    os.makedirs(os.path.dirname(out_xlsx) or ".", exist_ok=True)
    wb.save(out_xlsx)

    sheet_count = len(wb.sheetnames) - 1  # Trừ sheet Full Text
    print(f"✅ Excel  : {out_xlsx}  ({sheet_count} sheet bảng biểu + 1 sheet full text)")


# ──────────────────────────────────────────────────────────
# HELPER: Parse bảng từ Markdown (fallback khi docling
#         không expose tables riêng)
# ──────────────────────────────────────────────────────────
def _write_excel_from_markdown(wb, md_text: str):
    """Parse bảng Markdown dạng |..| và ghi vào wb."""
    lines = md_text.split("\n")
    current_table = []
    table_idx = 0

    def flush_table():
        nonlocal table_idx
        if len(current_table) < 2:
            return
        table_idx += 1
        ws = wb.create_sheet(title=f"Bảng {table_idx}")
        for row in current_table:
            ws.append([_parse_number(c.strip()) for c in row])
        apply_excel_style(ws, header_row_count=1)
        current_table.clear()

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            # Dòng separator (|---|---|) → bỏ qua
            inner = stripped[1:-1]
            cells = inner.split("|")
            if all(set(c.strip()) <= {"-", ":", " "} for c in cells):
                continue
            current_table.append(cells)
        else:
            if current_table:
                flush_table()

    if current_table:
        flush_table()

    if table_idx == 0:
        ws = wb.create_sheet(title="Nội dung")
        ws.append(["Không tìm thấy bảng biểu — xem sheet Full Text"])


def _parse_number(value):
    """Thử chuyển string thành số (int/float) nếu có thể."""
    if not isinstance(value, str):
        return value
    cleaned = value.replace(",", "").replace(" ", "").replace(".", "")
    # Thử int
    try:
        return int(cleaned)
    except ValueError:
        pass
    # Thử float (dấu phẩy thập phân VN)
    try:
        return float(value.replace(",", "."))
    except ValueError:
        pass
    return value


# ──────────────────────────────────────────────────────────
# BATCH: Convert toàn bộ folder
# ──────────────────────────────────────────────────────────
def batch_convert(folder_path: str, out_folder: str):
    """Batch convert tất cả PDF trong folder → mỗi file 1 cặp .md + .xlsx"""
    pdf_files = sorted(glob.glob(os.path.join(folder_path, "*.pdf")))

    if not pdf_files:
        print(f"⚠️  Không tìm thấy PDF nào trong: {folder_path}")
        return

    print(f"📂 Tìm thấy {len(pdf_files)} file PDF\n")
    os.makedirs(out_folder, exist_ok=True)

    success, fail = 0, 0
    for i, pdf_path in enumerate(pdf_files, 1):
        name = os.path.splitext(os.path.basename(pdf_path))[0]
        out_md   = os.path.join(out_folder, name + ".md")
        out_xlsx = os.path.join(out_folder, name + ".xlsx")

        print(f"\n⏳ ({i}/{len(pdf_files)}) {name[:65]}")
        print("─" * 60)
        try:
            convert_bctc(pdf_path, out_md, out_xlsx)
            success += 1
        except SystemExit:
            fail += 1
        except Exception as e:
            print(f"  ❌ Lỗi: {e}")
            fail += 1

    print(f"\n{'='*55}")
    print(f"=== KẾT QUẢ BATCH: {success} thành công, {fail} thất bại ===")
    print(f"{'='*55}")
    print(f"📁 Thư mục output: {out_folder}")


# ──────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="BCTC to Excel — PDF Báo cáo tài chính → Markdown + Excel",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "input",
        help="File PDF cần convert, hoặc FOLDER nếu dùng --batch",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="File .md xuất ra (mặc định: cùng tên PDF, đuôi .md)",
    )
    parser.add_argument(
        "--excel",
        default=None,
        help="File .xlsx xuất ra (mặc định: cùng tên PDF, đuôi .xlsx)",
    )
    parser.add_argument(
        "--batch",
        action="store_true",
        help="Batch convert toàn bộ folder (input là thư mục)",
    )
    parser.add_argument(
        "--outdir",
        default=None,
        help="Thư mục output khi dùng --batch (mặc định: [input]/BCTC Excel)",
    )

    args = parser.parse_args()

    # ── BATCH MODE ──
    if args.batch:
        if not os.path.isdir(args.input):
            print(f"❌ Lỗi: '{args.input}' không phải thư mục hợp lệ")
            sys.exit(1)
        out_folder = args.outdir or os.path.join(args.input, "BCTC Excel")
        batch_convert(args.input, out_folder)

    # ── SINGLE FILE MODE ──
    else:
        if not os.path.exists(args.input):
            print(f"❌ Lỗi: File không tồn tại — {args.input}")
            sys.exit(1)

        base = os.path.splitext(args.input)[0]
        out_md   = args.out   or (base + ".md")
        out_xlsx = args.excel or (base + ".xlsx")

        convert_bctc(args.input, out_md, out_xlsx)
